import numpy as np
import pandas as pd
import streamlit as st
import sys
import os
from os import listdir
from datetime import datetime
import pandas as pd
import re
import numpy as np
import re
import smtplib
import zlib
import zipfile
from email.mime.text import MIMEText


from openpyxl import load_workbook


def compress(file_paths):
	print("File Paths:")
	print(file_paths)

	# path = "C:/data/"

	# Select the compression mode ZIP_DEFLATED for compression
	# or zipfile.ZIP_STORED to just store the file
	compression = zipfile.ZIP_DEFLATED
	zip_path = "./dashboard_zip.zip"
	# create the zip file first parameter path/name, second mode
	zf = zipfile.ZipFile("./dashboard_zip.zip", mode="w")
	try:
		for file_path in file_paths:
			# Add file to the zip file
			# first parameter file to zip, second filename in zip
			zf.write(file_path, file_path, compress_type=compression)

	except FileNotFoundError:
		print("An error occurred")
	finally:
		# Don't forget to close the file!
		zf.close()
	return zip_path

# with st.echo():
st.markdown("[![Click me](./app/static/logo-2018-small.jpg)](https://www.second-opinions.org/home)")
st.title("Welcome to Second Opinion's Free clinic dashboard creator!")
def clear_sheet(sheet):
	for row in sheet:
		sheet.delete_rows(1, sheet.max_row+1)
	return sheet


def copy_paste_cleaned_data(wb, target_sheet_name, source_sheet, reporting_year):
	clear_sheet(wb[target_sheet_name])
	target_sheet = wb[target_sheet_name]

	# calculate total number of rows and 
	# columns in source excel file
	mr = source_sheet.max_row
	mc = source_sheet.max_column
	  
	# copying the cell values from source 
	# excel file to destination excel file
	for i in range (1, mr + 1):
		for j in range (1, mc + 1):
			# reading cell value from source excel file
			c = source_sheet.cell(row = i, column = j)
			# writing the read value to destination excel file
			target_sheet.cell(row = i, column = j).value = c.value
	target_sheet.insert_cols(0)
	target_sheet.insert_cols(0)
	target_sheet['B1'] = "Year"
	target_sheet['A1'] = "Helper"

	# create helper column and year column in dataset
	for i in range (2, mr + 1):
		if target_sheet.cell(row = i, column = 7).value == None:
			continue
		else:
			formula = "=G" + str(i) + "&B" + str(i) 
			target_sheet.cell(row = i, column = 1).value = formula
			target_sheet.cell(row = i, column = 2).value = reporting_year
	return wb

def convert_empty_to_zero(sheet, col):
	print(sheet.cell(row = 1, column = col).value)
	mr = sheet.max_row

	#start at row 2 to skip headers
	for i in range(2, mr + 1):  # Include mr in the range
		cell_value = sheet.cell(row=i, column=col).value

		print("the value was : " + str(cell_value))
		if isinstance(cell_value, str):
			# Trim and convert to lowercase for uniformity
			trimmed_value = cell_value.lower().strip()
			# Check if the string is "-" or "n/a"
			if trimmed_value in ["-", "n/a"]:
				sheet.cell(row=i, column=col).value = 0
			# Check if the string contains no numbers
			elif not any(char.isdigit() for char in trimmed_value):
				sheet.cell(row=i, column=col).value = 0
		elif cell_value is None:
			continue


	return sheet

# this function is designed to clean numerical columns that erroneously have strings in them
# for example: "Approximately 150" -> should be replaced with 150
#		 "Zero (0)" -> should be replaced with 0 
#		 "828 - DENTAL Imaging Only" -> replaced with 828
#		"-" -> replaced with 0 
def find_numbers(sheet, col):

	mr = sheet.max_row
	#start at row 2 to skip headers
	for i in range(2, mr):
		if sheet.cell(row = i, column = col).value == None:
			continue
		else:
			digits_found = re.findall(r'\d+', str(sheet.cell(row = i, column = col).value))
			if len(digits_found) >0:
				sheet.cell(row = i, column = col).value = int(digits_found[0])
			else:
				sheet.cell(row = i, column = col).value = 0
	return sheet

def clean_data(wb, data_sheet_name):

	numerical_col_list = [  "What was the total number of 30-day on-site prescriptions filled or medications dispensed by the clinic in the past year. Note that this is different from the number of medications prescribed by the clinic. Please provide your best estimate.",
							"What was your total cash-operating expenditure in the past year? ",
							"What is the approximate total revenue received from patient fees and reimbursements of services in the past year?",
							"What was your total cash-operating expenditure in the past year?",
							"What is the approximate total revenue received from patient fees and reimbursements of services in the past year?",   
							"Approximately how many volunteer-hours were provided at your clinic within the past year? Please provide your best estimate.",
							"Total number of NEW Patients in Past Year",
							"Total number of patients served in past year (new and established combined)",
							"Number of Dental VISITS",
							"Total Number of all Medical VISITS (both primary care AND specialty visits for both new and established patients)",
							"Number of Mental Health/Behavioral Health VISITS",
							"Total Patients VISITS (sum of above visits)",
							"Of the total MEDICAL visits, approximately what number of MEDICAL visits would have occurred at the ED if the clinic was not in operation? If an estimation cannot be provided, consider surveying patients this question during the visit.",
							"Of the total DENTAL visits, approximately what number of DENTAL visits would have occurred at the ED if the clinic was not in operation? If an estimation cannot be provided, consider surveying patients this question during the visit.", 
							"Number of In-House Imaging Tests in past year",	
							"Number of In-House Lab Tests in past year",
							"Number of In-House COVID Tests in past year",
							"Number of COVID Vaccinations in past year provided  in your clinic by an outside organization (e.g. the state or local public health department or a private provider)",
							"Number of COVID Vaccinations in past year provided independently by your clinic (vaccines administered by clinic personnel, not by an outside organization)",
							"% Diabetes Screening/Management",
							"% HTN Screening/Management",   
							"% Cancer Screening/Management",
							"% Obesity Screening/Management",   
							"% Dental Care",
							"% Sexual Health Screening/Management",
							"% Dyslipidemia/Hypercholesterolemia Screening/Management", 
							"% Mental Health Screening/management",
							"% Influenza Immunization",  
							"% Other Immunizations (ex: shingles, pneumonia, COVID, etc)",
							"% Asthma/COPD Management", 
							"% Dermatology Screening & Management", 
							"% Heart Disease Screening & Management",
							"% Vision Screenings and Exams",	
							"% Arthritis/Musculoskeletal Screening/Management", 
							"% Physicals (school, sport, or general)",
							"% Acute Injury Management",	
							"% Hearing Screening and Exam",
							"% Transgender FTM (female-to-male) Patients",
							"% Transgender MTF (male-to-female) Patients", 
							"% Gender Non-Conforming Patients",
							"% 0-17 years old", 
							"% 18-64 years old",
							"% 65+ years old",
							"% Latino or Hispanic",
							"% White",
							"% Black or African American",
							"% Asian",
							"% American Indian or Alaskan Native",
							"% Native Hawaiian or Pacific Islander",
							"% Multi-racial or Bi-Racial",
							"% Other",
							"% Unknown",   
							"% : Below 100% of FPL",
							"% : Between 100% and 200% of FPL",
							"% : Over 200% of FPL"] 
	ws = wb[data_sheet_name]
	column_headers_obs = ws[1]
	column_headers = []
	for col in column_headers_obs:
		if col.value != None:
			column_headers = column_headers + [col.value.strip()]
			print(col.value)
	mc = ws.max_column

	for col in numerical_col_list:
		col_index = column_headers.index(col.strip()) + 1
		print("The column was: " + col)
		print("The column from the list was: "  + column_headers[col_index])
		convert_empty_to_zero(ws,col_index)
		find_numbers(ws, col_index)
	return wb

def data_quality_check():
	return True


#sheet is the openpyxl worksheet with the raw data
def pull_clinic_list(raw_input_sheet, reporting_year):

	clinic_list = []
	for i, row in enumerate(raw_input_sheet):
		#skip if row is not the appropriate year
		if row[1].value != reporting_year:
			continue
		# Skip the first row (the row with the column names)
		if i == 0:
			continue
		if row[0].value == None:
			continue

		else:
		# Get the value of the first cell in the row
			clinic_name = row[6].value
		# Add the value to the list
			clinic_list.append(clinic_name)
			print(clinic_name)
	return clinic_list

def update_clinic_dashbaords(dest,updated_dest, clinic_name, reporting_year):
	#Open an xlsx for reading
	wb = load_workbook(filename = dest)
	#Get the current Active Sheet
	ws = wb['0. Dashboard']

	#You can also select a particular sheet
	#based on sheet name
	#ws = wb.get_sheet_by_name("Sheet1")
	#Open the csv file
	ws['I12'] = reporting_year 
	ws['R12'] = clinic_name
	wb.save(updated_dest)

def check_password():
	"""Returns `True` if the user had the correct password."""

	def password_entered():
		"""Checks whether a password entered by the user is correct."""
		if st.session_state["password"] == st.secrets["password"]:
			st.session_state["password_correct"] = True
			del st.session_state["password"]  # don't store password
		else:
			st.session_state["password_correct"] = False

	if "password_correct" not in st.session_state:
		# First run, show input for password.
		st.text_input(
			"Password", type="password", on_change=password_entered, key="password"
		)
		return False
	elif not st.session_state["password_correct"]:
		# Password not correct, show input + error.
		st.text_input(
			"Password", type="password", on_change=password_entered, key="password"
		)
		st.error("😕 Password incorrect")
		return False
	else:
		# Password correct.
		return True


def copy_paste_conversions(target_wb, source_sheet):
	reimbursment_sheet = target_wb['1. Reimbursement Value']
	ed_sheet = target_wb['3. Diverted ED Health']
	reimbursment_sheet['C29'] = source_sheet['D5'].value
	reimbursment_sheet['C30'] = source_sheet['D6'].value	
	reimbursment_sheet['C31'] = source_sheet['D7'].value
	reimbursment_sheet['C33'] = source_sheet['D9'].value
	reimbursment_sheet['C35'] = source_sheet['D11'].value
	reimbursment_sheet['C37'] = source_sheet['D13'].value
	reimbursment_sheet['C39'] = source_sheet['D15'].value
	reimbursment_sheet['C40'] = source_sheet['D16'].value
	reimbursment_sheet['C41'] = source_sheet['D17'].value
	reimbursment_sheet['C43'] = source_sheet['D19'].value
	reimbursment_sheet['C45'] = source_sheet['D21'].value
	reimbursment_sheet['C47'] = source_sheet['D23'].value
	ed_sheet['D23'] = source_sheet['D27'].value
	ed_sheet['D28'] = source_sheet['D32'].value
	ed_sheet['D30'] = source_sheet['D34'].value
	return wb

def click_button():
	st.session_state.button = not st.session_state.button




if check_password():
	if 'button' not in st.session_state:
		st.session_state.button = False

	conversion_template_file_path = "./static/fcc-dashboard-conversion-template.xlsx"
	st.subheader('You can use default conversion rates or upload your own custom conversions.')
	st.subheader('You can download a template conversion file here')
	with open(conversion_template_file_path, 'rb') as f:
		st.download_button('Download', f, file_name="fcc-dashboard-conversion-template.xlsx") 

	st.button('Toggle Custom Conversion', on_click=click_button)

	if st.session_state.button:
		# The message and nested widget will remain on the page
		st.write('Custom conversion on')
	else:
		st.write('Custom conversion off')
	dest = "./static/2022-IAFCC-Master-Dashboard-source.xlsx"
	default_conversion_path = "./static/fcc-dashboard-default-conversions.xlsx"
	updated_dest = "./static/Overall-Clinic-Dashboard.xlsx"
	dashboard_folder_path = "./static/"

	st.text("")
	spectra = st.file_uploader("Upload your clinic's data here! (*max 1000 clinics*)", type={"csv", "xlsx"})
	reporting_year = 2022
	if spectra is not None:
		
		if st.session_state.button:
			conversions = st.file_uploader("Upload your conversion template here! (*max 1000 clinics*)", type={"csv", "xlsx"})
		else:
			conversions = default_conversion_path
		if conversions is not None:
			# conversions values
			conversions_wb = load_workbook(conversions)
			conversions_ws = conversions_wb.active

			# clinic data sheets
			clean_source_wb = load_workbook(spectra)
			clean_source_ws = clean_source_wb.active
			st.write("Completed upload")
			wb = load_workbook(filename = dest)
			wb = copy_paste_cleaned_data(wb, 'Cleaned Responses',clean_source_ws, reporting_year = reporting_year)
			wb = copy_paste_conversions(wb, conversions_ws)
			wb = clean_data(wb,"Cleaned Responses")
			clinic_list = pull_clinic_list(wb['Cleaned Responses'], reporting_year)
			# wb['Raw Model Inputs'].delete_columns(1,1)
			mr = wb['Raw Model Inputs'].max_row
			# remove unneccesary rows
			# add in appropriate data from uploaded sheet
			# for loop must go in reverse to prevent skipping rows in excel (as index moves when you remove rows)
			for i in range(1000,-1,-1):

				if i <len(clinic_list):
					print(clinic_list[i])
					wb['Raw Model Inputs'].cell(row = i+2, column = 1).value = clinic_list[i]
					wb['Raw Model Inputs'].cell(row = i+2, column = 2).value = reporting_year
				else: 
					wb['Raw Model Inputs'].delete_rows(idx=i+2)
					
			wb['0. Dashboard']["I12"].value = reporting_year
			wb.save(updated_dest)

			file_paths = [updated_dest]
			for clinic in clinic_list:
				final_path = dashboard_folder_path + str(reporting_year) + "-" + clinic + " dashboard.xlsx"
				file_paths = file_paths + [final_path]
				update_clinic_dashbaords(updated_dest, final_path, clinic, reporting_year)
				
			zip_path = compress(file_paths)
			with open(zip_path, 'rb') as f:
				st.text("Download All Dashboards Here:")
				st.download_button('Download ZIP', f, file_name= 'dashboards.zip')  # Defaults to 'application/octet-stream'

			for file_path in file_paths:
				with open(file_path, 'rb') as f:
					dashboard_name = file_path.split('/')[-1]
					clinic_name = dashboard_name.split('.')[0]
					print(dashboard_name)
					st.text(str(clinic_name))
					st.download_button('Download', f, file_name= str(dashboard_name))  # Defaults to 'application/octet-stream'

